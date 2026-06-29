"""
Excel ve PDF rapor üretim servisi.
openpyxl (Excel) ve reportlab (PDF) kullanır.
"""
from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db.models import DecimalField, Sum
from django.db.models.functions import Coalesce

from apps.budget.models import Budget
from apps.projects.models import Project


# ---------------------------------------------------------------------------
# Yardımcı: proje verisi çek
# ---------------------------------------------------------------------------

def _get_project_rows(project_id=None) -> list[dict]:
    """
    Proje + bütçe verilerini tek sorgu grubunda çeker.
    Her satır bir dict döner.
    """
    money = DecimalField(max_digits=18, decimal_places=2)

    qs = Project.objects.select_related("budget").order_by("province", "name")
    if project_id is not None:
        qs = qs.filter(pk=project_id)

    # Bütçe toplamlarını tek annotate ile almak için Budget üzerinden gidiyoruz
    budget_map: dict[int, dict] = {}
    budgets = Budget.objects.annotate(
        total_spent=Coalesce(Sum("expenses__amount"), Decimal("0"), output_field=money)
    ).select_related("project")
    if project_id is not None:
        budgets = budgets.filter(project_id=project_id)
    for b in budgets:
        budget_map[b.project_id] = {
            "planned": b.planned_amount,
            "spent": b.total_spent,
            "remaining": b.planned_amount - b.total_spent,
        }

    rows = []
    for p in qs:
        bdata = budget_map.get(p.id, {"planned": Decimal("0"), "spent": Decimal("0"), "remaining": Decimal("0")})
        rows.append({
            "name": p.name,
            "province": p.province,
            "status": p.get_status_display(),
            "progress": p.progress,
            "planned_end": p.planned_end,
            "actual_end": p.actual_end,
            "is_delayed": p.is_delayed,
            "planned": bdata["planned"],
            "spent": bdata["spent"],
            "remaining": bdata["remaining"],
        })
    return rows


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def generate_summary_excel(project_id=None) -> bytes:
    """
    openpyxl ile Excel dosyası üretir; bytes döner.
    project_id verilirse tek proje, verilmezse tüm projeler raporu.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Proje Raporu"

    # Başlık
    title_text = "ProjeHaritası — Proje Raporu"
    ws.merge_cells("A1:J1")
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="4f6ef7")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Oluşturulma: {date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = Font(italic=True, color="888888")
    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 18

    # Sütun başlıkları
    headers = [
        "Proje Adı",
        "İl",
        "Durum",
        "İlerleme (%)",
        "Plan. Bitiş",
        "Gerçek Bitiş",
        "Gecikiyor",
        "Bütçe (TL)",
        "Harcama (TL)",
        "Kalan (TL)",
    ]
    header_fill = PatternFill("solid", fgColor="4f6ef7")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 20

    # Veri satırları
    rows = _get_project_rows(project_id)
    number_fmt = "#,##0.00"

    for row_idx, r in enumerate(rows, start=4):
        ws.cell(row=row_idx, column=1, value=r["name"])
        ws.cell(row=row_idx, column=2, value=r["province"])
        ws.cell(row=row_idx, column=3, value=r["status"])
        ws.cell(row=row_idx, column=4, value=r["progress"])
        ws.cell(row=row_idx, column=5, value=r["planned_end"].strftime("%d.%m.%Y") if r["planned_end"] else "")
        ws.cell(row=row_idx, column=6, value=r["actual_end"].strftime("%d.%m.%Y") if r["actual_end"] else "")
        ws.cell(row=row_idx, column=7, value="Evet" if r["is_delayed"] else "Hayır")

        budget_cell = ws.cell(row=row_idx, column=8, value=float(r["planned"]))
        budget_cell.number_format = number_fmt

        spent_cell = ws.cell(row=row_idx, column=9, value=float(r["spent"]))
        spent_cell.number_format = number_fmt

        remaining_cell = ws.cell(row=row_idx, column=10, value=float(r["remaining"]))
        remaining_cell.number_format = number_fmt

    # Sütun genişlikleri
    col_widths = [40, 15, 14, 14, 14, 14, 12, 18, 18, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def generate_summary_pdf(project_id=None) -> bytes:
    """
    reportlab SimpleDocTemplate ile PDF üretir; bytes döner.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=4,
        textColor=colors.HexColor("#4f6ef7"),
    )
    sub_style = ParagraphStyle(
        "ReportSub",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.HexColor("#888888"),
        spaceAfter=10,
    )

    story = [
        Paragraph("ProjeHaritası — Proje Raporu", title_style),
        Paragraph(f"Oluşturulma tarihi: {date.today().strftime('%d.%m.%Y')}", sub_style),
        Spacer(1, 0.3 * cm),
    ]

    # Tablo verisi
    rows = _get_project_rows(project_id)

    header_row = [
        "Proje Adı",
        "İl",
        "Durum",
        "İlerleme\n(%)",
        "Plan. Bitiş",
        "Gerçek Bitiş",
        "Gecikiyor",
        "Bütçe (TL)",
        "Harcama (TL)",
        "Kalan (TL)",
    ]

    table_data = [header_row]
    alt_color = colors.HexColor("#f0f4ff")
    header_bg = colors.HexColor("#4f6ef7")

    for r in rows:
        table_data.append([
            r["name"],
            r["province"],
            r["status"],
            str(r["progress"]),
            r["planned_end"].strftime("%d.%m.%Y") if r["planned_end"] else "",
            r["actual_end"].strftime("%d.%m.%Y") if r["actual_end"] else "",
            "Evet" if r["is_delayed"] else "Hayır",
            f"{float(r['planned']):,.2f}",
            f"{float(r['spent']):,.2f}",
            f"{float(r['remaining']):,.2f}",
        ])

    # Sütun genişlikleri (toplam A4 yatay ~25.7 cm)
    col_widths_cm = [5.5, 2.2, 2.4, 1.6, 2.3, 2.3, 1.8, 2.8, 2.8, 2.8]
    col_widths_pt = [w * cm for w in col_widths_cm]

    tbl = Table(table_data, colWidths=col_widths_pt, repeatRows=1)

    style_cmds = [
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        # Veri satırları
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 1), (2, -1), "LEFT"),
        ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, alt_color]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    tbl.setStyle(TableStyle(style_cmds))

    story.append(tbl)
    doc.build(story)
    return buf.getvalue()
